"""
Generatore di file xlsx compatibile con il formato import/export PriMus.

Struttura per ogni voce (colonne A→N, 1-based):
  A: vuota (spacing)
  B: Nr.Ord.  C: TARIFFA  D: DESIGNAZIONE  E: Par.ug  F: Lung.  G: Larg.  H: H/peso
  I: Quantità (formula ROUND/PRODUCT o valore diretto)
  J: Prezzo unitario (prezzo finito con ricarica)
  K: TOTALE (formula ROUND/PRODUCT)
  L, M, N: ClDes, ClQT, Linha (metadati nascosti PriMus)

Sequenza righe per ciascuna voce:
  1. Intestazione voce   → B=nr, C=codice, D=descrizione
  2. Label misurazioni   → D='M I S U R A Z I O N I:'
  3..n. Righe misurazioni → D=desc, E-H=fattori, I=ROUND(PRODUCT(...),2)
  n+1. Riga marker       → N="3'"
  n+2. Riga SOMMANO      → D='SOMMANO [um]', I=ROUND(SUM(...),2), J=prezzo, K=ROUND(PRODUCT(...),2)
  n+3. Riga separatore   → D=''

Righe finali:
  TOTALE euro           → D='TOTALE euro', K=ROUND(SUM(K4:Klast),2), L=M=N='0'
  (vuota)
  AGGIUNGE NUOVA VOCE   → D='AGGIUNGE NUOVA VOCE', L=M=N='0'
  (vuota)
  Footer                → B='documento realizzato conPriMus for Excel'

Note tecniche:
  - openpyxl scrive stringhe come inline strings (t="inlineStr"); PriMus
    richiede shared strings (t="s"). La funzione ``_converti_in_shared_strings``
    post-processa il file xlsx convertendo il formato.
  - Le celle «placeholder» (senza formula reale) vengono lasciate vuote
    anziché usare <f>=</f>, che è invalido per Excel.
"""

from __future__ import annotations
import io
import re
import zipfile

from openpyxl import Workbook
from openpyxl.cell import Cell

from cbd_preventivi.models import Preventivo
from cbd_preventivi.calcoli import prezzo_per_um


# Indici colonna (1-based)
COL_A = 1
COL_B, COL_C, COL_D = 2, 3, 4
COL_E, COL_F, COL_G, COL_H = 5, 6, 7, 8
COL_I, COL_J, COL_K = 9, 10, 11
COL_L, COL_M, COL_N = 12, 13, 14

# Prima riga dati (le righe 2 e 3 contengono le intestazioni di colonna)
PRIMA_RIGA_DATI = 4


# ---------------------------------------------------------------------------
# Helpers di scrittura celle
# ---------------------------------------------------------------------------

def _cella(foglio, riga: int, colonna: int) -> Cell:
    return foglio.cell(row=riga, column=colonna)


def _scrivi_formula(foglio, riga: int, colonna: int, formula: str) -> None:
    """Scrive una formula Excel nella cella specificata."""
    cella = foglio.cell(row=riga, column=colonna)
    cella.value = formula
    cella.data_type = "f"


# ---------------------------------------------------------------------------
# Costruzione foglio
# ---------------------------------------------------------------------------

def _scrivi_intestazioni(foglio) -> None:
    """Righe 2 e 3: intestazione delle colonne del computo."""
    foglio.cell(row=2, column=COL_B).value = "Nr. Ord."
    foglio.cell(row=2, column=COL_C).value = "TARIFFA"
    foglio.cell(row=2, column=COL_D).value = "DESIGNAZIONE DEI LAVORI"
    foglio.cell(row=2, column=COL_F).value = "M I S U R A Z I O N I:"
    foglio.cell(row=2, column=COL_I).value = "Quantità"
    foglio.cell(row=2, column=COL_J).value = "         IMPORTI"

    foglio.cell(row=3, column=COL_B).value = " "
    foglio.cell(row=3, column=COL_C).value = "  "
    foglio.cell(row=3, column=COL_D).value = "   "
    foglio.cell(row=3, column=COL_E).value = "Par.ug"
    foglio.cell(row=3, column=COL_F).value = "Lung."
    foglio.cell(row=3, column=COL_G).value = "Larg."
    foglio.cell(row=3, column=COL_H).value = "H/peso"
    foglio.cell(row=3, column=COL_I).value = "    "
    foglio.cell(row=3, column=COL_J).value = "unitario"
    foglio.cell(row=3, column=COL_K).value = "TOTALE"
    foglio.cell(row=3, column=COL_L).value = "ClDes"
    foglio.cell(row=3, column=COL_M).value = "ClQT"
    foglio.cell(row=3, column=COL_N).value = "Linha"


def _imposta_colonne(foglio) -> None:
    """Larghezze e visibilità delle colonne come nel file PriMus originale."""
    dimensioni = {
        "A": (2.33,  False),
        "B": (5.5,   False),
        "C": (13.5,  False),
        "D": (55.5,  False),
        "E": (10.83, False),
        "J": (12.0,  False),
        "K": (16.0,  False),
        "L": (8.5,   True),
        "M": (7.5,   True),
        "N": (7.16,  True),
    }
    for lettera, (larghezza, nascosta) in dimensioni.items():
        foglio.column_dimensions[lettera].width = larghezza
        foglio.column_dimensions[lettera].hidden = nascosta


def _scrivi_voce(foglio, indice_voce: int, voce, prezzo_unitario: float, riga_corrente: int) -> int:
    """Scrive tutte le righe di una voce a partire da ``riga_corrente``.

    Returns:
        Numero della prima riga disponibile dopo il separatore della voce.
    """
    numero_voce = str(indice_voce + 1)
    riga_intestazione = riga_corrente

    # 1. Riga intestazione voce
    _cella(foglio, riga_intestazione, COL_B).value = numero_voce
    _cella(foglio, riga_intestazione, COL_C).value = voce.codice
    _cella(foglio, riga_intestazione, COL_D).value = voce.descrizione

    # 2. Riga label «M I S U R A Z I O N I:»
    riga_label_mis = riga_intestazione + 1
    _cella(foglio, riga_label_mis, COL_D).value = "M I S U R A Z I O N I:"

    # 3. Righe misurazioni
    misurazioni_valide = [m for m in voce.misurazioni if not m.is_empty]

    if misurazioni_valide:
        righe_da_scrivere = misurazioni_valide
        usa_quantita_manuale = False
    else:
        # Nessuna misurazione: placeholder con quantita_manuale in par_ug
        righe_da_scrivere = [None]
        usa_quantita_manuale = True

    riga_prima_mis = riga_label_mis + 1
    ultima_riga_mis = riga_prima_mis

    for offset, misurazione in enumerate(righe_da_scrivere):
        riga = riga_prima_mis + offset
        ultima_riga_mis = riga

        if misurazione is not None:
            _cella(foglio, riga, COL_D).value = misurazione.descrizione

            if misurazione.quantita_diretta is not None:
                # Quantità diretta: valore numerico in I, E-H vuoti.
                # Così il round-trip import→export→import preserva quantita_diretta.
                _cella(foglio, riga, COL_I).value = misurazione.quantita_diretta
            else:
                if misurazione.par_ug is not None:
                    _cella(foglio, riga, COL_E).value = misurazione.par_ug
                if misurazione.lung is not None:
                    _cella(foglio, riga, COL_F).value = misurazione.lung
                if misurazione.larg is not None:
                    _cella(foglio, riga, COL_G).value = misurazione.larg
                if misurazione.h_peso is not None:
                    _cella(foglio, riga, COL_H).value = misurazione.h_peso
                _scrivi_formula(foglio, riga, COL_I, f"=ROUND(PRODUCT(E{riga}:H{riga}),2)")

        elif usa_quantita_manuale and voce.quantita_manuale is not None:
            _cella(foglio, riga, COL_E).value = voce.quantita_manuale
            _scrivi_formula(foglio, riga, COL_I, f"=ROUND(PRODUCT(E{riga}:H{riga}),2)")
        else:
            _scrivi_formula(foglio, riga, COL_I, f"=ROUND(PRODUCT(E{riga}:H{riga}),2)")

    # 4. Riga marker «3'»
    riga_marker = ultima_riga_mis + 1
    _cella(foglio, riga_marker, COL_N).value = "3'"

    # 5. Riga SOMMANO
    riga_sommano = riga_marker + 1
    _cella(foglio, riga_sommano, COL_D).value = f"SOMMANO {voce.um}"
    _scrivi_formula(foglio, riga_sommano, COL_I,
                    f"=ROUND(SUM(I{riga_label_mis}:I{riga_marker}),2)")
    _cella(foglio, riga_sommano, COL_J).value = prezzo_unitario
    _scrivi_formula(foglio, riga_sommano, COL_K,
                    f"=ROUND(PRODUCT(I{riga_sommano}:J{riga_sommano}),2)")

    # 6. Riga separatore
    riga_separatore = riga_sommano + 1
    _cella(foglio, riga_separatore, COL_D).value = ""

    return riga_separatore + 1


def _scrivi_footer(foglio, riga_dopo_ultimo_separatore: int) -> None:
    """Righe finali: TOTALE euro, AGGIUNGE NUOVA VOCE, firma PriMus."""
    riga_totale = riga_dopo_ultimo_separatore
    _cella(foglio, riga_totale, COL_D).value = "TOTALE euro"
    _scrivi_formula(foglio, riga_totale, COL_K,
                    f"=ROUND(SUM(K{PRIMA_RIGA_DATI}:K{riga_totale - 1}),2)")
    _cella(foglio, riga_totale, COL_L).value = "0"
    _cella(foglio, riga_totale, COL_M).value = "0"
    _cella(foglio, riga_totale, COL_N).value = "0"

    riga_nuova_voce = riga_totale + 2
    _cella(foglio, riga_nuova_voce, COL_D).value = "AGGIUNGE NUOVA VOCE"
    _cella(foglio, riga_nuova_voce, COL_L).value = "0"
    _cella(foglio, riga_nuova_voce, COL_M).value = "0"
    _cella(foglio, riga_nuova_voce, COL_N).value = "0"

    riga_firma = riga_nuova_voce + 2
    _cella(foglio, riga_firma, COL_B).value = "documento realizzato conPriMus for Excel"


# ---------------------------------------------------------------------------
# Post-processing: inline strings → shared strings
# ---------------------------------------------------------------------------

def _converti_in_shared_strings(xlsx_bytes: bytes) -> bytes:
    """Converte le inline strings di openpyxl in shared strings.

    PriMus richiede ``xl/sharedStrings.xml`` e celle con ``t="s"``; openpyxl
    scrive ``t="inlineStr"``. Questa funzione post-processa lo zip dell'xlsx
    aggiungendo il file e aggiornando i riferimenti nelle celle.
    """
    zip_input = zipfile.ZipFile(io.BytesIO(xlsx_bytes))
    nomi_fogli = sorted(
        n for n in zip_input.namelist()
        if re.match(r"xl/worksheets/sheet\d+\.xml$", n)
    )

    # Raccoglie tutte le stringhe uniche (già XML-escaped) dai worksheet
    stringhe: list[str] = []
    indice_stringa: dict[str, int] = {}
    contenuti_fogli: dict[str, str] = {}

    for nome in nomi_fogli:
        contenuto = zip_input.read(nome).decode("utf-8")
        contenuti_fogli[nome] = contenuto
        for match in re.finditer(r"<is><t(?:[^>]*)>(.*?)</t></is>", contenuto, re.DOTALL):
            stringa = match.group(1)
            if stringa not in indice_stringa:
                indice_stringa[stringa] = len(stringhe)
                stringhe.append(stringa)

    def converti_foglio(contenuto: str) -> str:
        """Sostituisce le inline string con riferimenti a shared strings."""
        def sostituisci_stringa(match):
            pre, post, testo = match.group(1), match.group(2), match.group(3)
            return f'<c {pre}t="s"{post}><v>{indice_stringa[testo]}</v></c>'

        # Celle con stringa non vuota
        contenuto = re.sub(
            r'<c ([^>]*?)t="inlineStr"([^>]*)><is><t(?:[^>]*)>(.*?)</t></is></c>',
            sostituisci_stringa,
            contenuto,
            flags=re.DOTALL,
        )
        # Celle stringa vuota (self-closing) → rimuovi (cella assente = vuota)
        contenuto = re.sub(r'<c [^>]*t="inlineStr"\s*/>', "", contenuto)
        # Celle con formula vuota → celle vuote (evita <f>=</f> invalido per Excel)
        contenuto = re.sub(r"<c ([^>]*)>\s*<f\s*/>\s*<v\s*/>\s*</c>", r"<c \1/>", contenuto)
        return contenuto

    # Costruisce xl/sharedStrings.xml
    parti_shared_strings = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        f' count="{len(stringhe)}" uniqueCount="{len(stringhe)}">',
    ]
    for stringa in stringhe:
        preserve = 'xml:space="preserve"' if (stringa.startswith(" ") or stringa.endswith(" ")) else ""
        parti_shared_strings.append(
            f"<si><t {preserve}>{stringa}</t></si>" if preserve else f"<si><t>{stringa}</t></si>"
        )
    parti_shared_strings.append("</sst>")

    # Aggiorna [Content_Types].xml
    content_types = zip_input.read("[Content_Types].xml").decode("utf-8")
    content_types = content_types.replace(
        "</Types>",
        '<Override PartName="/xl/sharedStrings.xml"'
        ' ContentType="application/vnd.openxmlformats-officedocument'
        '.spreadsheetml.sharedStrings+xml"/></Types>',
    )

    # Aggiorna xl/_rels/workbook.xml.rels
    relazioni = zip_input.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    relazioni = relazioni.replace(
        "</Relationships>",
        '<Relationship Id="rIdSS"'
        ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings"'
        ' Target="sharedStrings.xml"/></Relationships>',
    )

    # Ricostruisce lo zip
    buffer_output = io.BytesIO()
    with zipfile.ZipFile(buffer_output, "w", zipfile.ZIP_DEFLATED) as zip_output:
        for nome in zip_input.namelist():
            if nome in contenuti_fogli:
                zip_output.writestr(nome, converti_foglio(contenuti_fogli[nome]))
            elif nome == "[Content_Types].xml":
                zip_output.writestr(nome, content_types)
            elif nome == "xl/_rels/workbook.xml.rels":
                zip_output.writestr(nome, relazioni)
            else:
                zip_output.writestr(nome, zip_input.read(nome))
        zip_output.writestr("xl/sharedStrings.xml", "".join(parti_shared_strings))

    buffer_output.seek(0)
    return buffer_output.read()


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def genera_xlsx(preventivo: Preventivo) -> bytes:
    """Genera il file xlsx nel formato PriMus e restituisce i bytes pronti per il download.

    Args:
        preventivo: il ``Preventivo`` da esportare.

    Returns:
        Bytes del file xlsx compatibile con PriMus ed Excel.
    """
    workbook = Workbook()
    foglio = workbook.active
    foglio.title = "Computo metrico"

    _scrivi_intestazioni(foglio)
    _imposta_colonne(foglio)

    riga_corrente = PRIMA_RIGA_DATI
    for indice, voce in enumerate(preventivo.voci):
        prezzo = prezzo_per_um(voce, preventivo)
        riga_corrente = _scrivi_voce(foglio, indice, voce, prezzo, riga_corrente)

    _scrivi_footer(foglio, riga_corrente)

    # Foglio "Dati" vuoto (presente nel formato PriMus originale)
    workbook.create_sheet("Dati")

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return _converti_in_shared_strings(buffer.read())
