"""
Parser per file xlsx nel formato import/export PriMus.

Legge un file xlsx esportato da PriMus e restituisce un ``Preventivo``
con le voci e le misurazioni corrispondenti.

Struttura attesa del foglio (colonne A→N, 1-based):
  B: Nr.Ord.   C: TARIFFA   D: DESIGNAZIONE
  E: Par.ug    F: Lung.     G: Larg.    H: H/peso
  I: Quantità  J: Prezzo unitario       N: Linha (marker)

Il parser usa una macchina a stati per riconoscere le sezioni:
  seek_voce → in_header → in_mis → after_marker → after_sommano
"""

from __future__ import annotations
import io
from typing import Optional

from openpyxl import load_workbook

from cbd_preventivi.models import Preventivo, Voce, RigaMisurazione


# Indici colonna (1-based)
COL_B, COL_C, COL_D = 2, 3, 4
COL_E, COL_F, COL_G, COL_H = 5, 6, 7, 8
COL_I, COL_J = 9, 10
COL_N = 14

PRIMA_RIGA_DATI = 4

# Mappa delle unità di misura PriMus → etichette standard del tool
_MAPPA_UM: dict[str, str] = {
    "a":       "a corpo",
    "cadauno": "cad",
    "m2":      "mq",
    "m3":      "mc",
}


def _a_float(valore) -> Optional[float]:
    """Converte un valore di cella in float; restituisce None se non numerico."""
    if valore is None:
        return None
    try:
        return float(valore)
    except (TypeError, ValueError):
        return None


def _a_stringa(valore) -> str:
    """Converte un valore di cella in stringa pulita."""
    return str(valore).strip() if valore is not None else ""


def _leggi_sommano(voce_corrente: dict, testo_d: str, valore_j) -> None:
    """Estrae UM e costo unitario dalla riga SOMMANO.

    Il testo in D è del tipo «SOMMANO mq»; il valore in J è il costo/UM
    così come appare in PriMus (senza ricarica).
    """
    parti = testo_d.split()
    if len(parti) > 1:
        um_primus = parti[1].lower()
        voce_corrente["um"] = _MAPPA_UM.get(um_primus, parti[1])
    costo = _a_float(valore_j)
    if costo is not None:
        voce_corrente["costo_override"] = costo


def _salva_voce(lista_voci: list[Voce], voce_corrente: dict) -> None:
    """Finalizza la voce corrente e la aggiunge alla lista.

    Se c'è una sola riga di misurazione senza descrizione e solo par_ug
    valorizzato, viene interpretata come segnaposto per ``quantita_manuale``
    (il modo in cui PriMus rappresenta le voci «a corpo» senza misure esplicite).
    """
    misurazioni: list[RigaMisurazione] = voce_corrente.get("misurazioni", [])
    quantita_manuale: Optional[float] = None

    # Rileva il segnaposto quantita_manuale: una sola riga, nessuna descrizione,
    # solo par_ug impostato, tutti gli altri fattori assenti
    if len(misurazioni) == 1 and not misurazioni[0].descrizione.strip():
        riga = misurazioni[0]
        if (riga.par_ug is not None
                and riga.lung is None
                and riga.larg is None
                and riga.h_peso is None):
            quantita_manuale = riga.par_ug
            misurazioni = []

    lista_voci.append(Voce(
        codice=voce_corrente["codice"],
        descrizione=voce_corrente["descrizione"],
        um=voce_corrente.get("um", ""),
        costo_override=voce_corrente.get("costo_override"),
        quantita_manuale=quantita_manuale,
        misurazioni=misurazioni,
    ))


def parse_primus_xlsx(contenuto_file: bytes) -> Preventivo:
    """Legge un file xlsx PriMus e restituisce un ``Preventivo`` con le voci importate.

    Args:
        contenuto_file: bytes del file xlsx.

    Returns:
        Un ``Preventivo`` con le voci estratte dal file.
    """
    workbook = load_workbook(io.BytesIO(contenuto_file), data_only=True)
    foglio = workbook.active

    lista_voci: list[Voce] = []
    stato = "seek_voce"
    voce_corrente: Optional[dict] = None

    for numero_riga in range(PRIMA_RIGA_DATI, foglio.max_row + 1):
        val_b = foglio.cell(numero_riga, COL_B).value
        val_c = foglio.cell(numero_riga, COL_C).value
        val_d = foglio.cell(numero_riga, COL_D).value
        val_e = foglio.cell(numero_riga, COL_E).value
        val_f = foglio.cell(numero_riga, COL_F).value
        val_g = foglio.cell(numero_riga, COL_G).value
        val_h = foglio.cell(numero_riga, COL_H).value
        val_i = foglio.cell(numero_riga, COL_I).value
        val_j = foglio.cell(numero_riga, COL_J).value
        val_n = foglio.cell(numero_riga, COL_N).value

        testo_d = _a_stringa(val_d)
        testo_b = _a_stringa(val_b)

        # Riga di footer → fine parsing
        if (testo_d in ("TOTALE euro", "AGGIUNGE NUOVA VOCE")
                or testo_b == "documento realizzato conPriMus for Excel"):
            if voce_corrente:
                _salva_voce(lista_voci, voce_corrente)
            break

        if stato == "seek_voce":
            if testo_b.isdigit() and int(testo_b) > 0:
                voce_corrente = {
                    "codice": _a_stringa(val_c),
                    "descrizione": testo_d,
                    "um": "",
                    "costo_override": None,
                    "misurazioni": [],
                }
                stato = "in_header"

        elif stato == "in_header":
            if testo_d == "M I S U R A Z I O N I:":
                stato = "in_mis"
            elif testo_d.startswith("SOMMANO"):
                _leggi_sommano(voce_corrente, testo_d, val_j)
                stato = "after_sommano"

        elif stato == "in_mis":
            if _a_stringa(val_n) == "3'":
                stato = "after_marker"
            elif testo_d.startswith("SOMMANO"):
                _leggi_sommano(voce_corrente, testo_d, val_j)
                stato = "after_sommano"
            else:
                par_ug = _a_float(val_e)
                lung   = _a_float(val_f)
                larg   = _a_float(val_g)
                h_peso = _a_float(val_h)
                quantita_diretta = None

                # Riga «Vedi voce n° X»: E-H vuoti, quantità già calcolata in I (cache)
                if par_ug is None and lung is None and larg is None and h_peso is None:
                    i_cache = _a_float(val_i)
                    if i_cache:
                        quantita_diretta = i_cache

                riga = RigaMisurazione(
                    descrizione=testo_d,
                    par_ug=par_ug,
                    lung=lung,
                    larg=larg,
                    h_peso=h_peso,
                    quantita_diretta=quantita_diretta,
                )
                if not riga.is_empty or testo_d:
                    voce_corrente["misurazioni"].append(riga)

        elif stato == "after_marker":
            if testo_d.startswith("SOMMANO"):
                _leggi_sommano(voce_corrente, testo_d, val_j)
                stato = "after_sommano"

        elif stato == "after_sommano":
            # La riga separatore chiude la voce
            _salva_voce(lista_voci, voce_corrente)
            voce_corrente = None
            stato = "seek_voce"

    if voce_corrente:
        _salva_voce(lista_voci, voce_corrente)

    return Preventivo(voci=lista_voci)
