"""
Test sulla struttura dell'xlsx generato da ``primus.export``.

Verifica che il file prodotto rispetti il formato PriMus riga per riga,
usando un preventivo con le prime 3 voci reali del file
«cOMPUTO MIGLIO EXCEL PROVA.xlsx» come riferimento.
"""

import io

import openpyxl
import pytest

from cbd_preventivi.models import Preventivo, Voce, RigaMisurazione, RigaCosto
from cbd_preventivi.primus.export import genera_xlsx, PRIMA_RIGA_DATI
from cbd_preventivi.calcoli import quantita_x, prezzo_per_um


# ---------------------------------------------------------------------------
# Dataset di riferimento
# ---------------------------------------------------------------------------

def _preventivo_di_test() -> Preventivo:
    """Preventivo con le prime 3 voci reali del file PriMus di esempio."""
    return Preventivo(
        nome="Test PriMus",
        ricarica_default=0.0,
        voci=[
            Voce(
                codice="001",
                descrizione="Approntamento area di cantiere e oneri accessori",
                um="a corpo",
                misurazioni=[RigaMisurazione(par_ug=1.0)],
            ),
            Voce(
                codice="002",
                descrizione="Stripe out impiantistico elettrico, idraulico e termico",
                um="a corpo",
                misurazioni=[RigaMisurazione(par_ug=1.0)],
            ),
            Voce(
                codice="003",
                descrizione="Demolizione manuale delle tramezzature esistenti",
                um="mc",
                misurazioni=[
                    RigaMisurazione(descrizione="Tramezza sp 20 esistente soggiorno",
                                    lung=4.2, larg=0.2, h_peso=3.3),
                    RigaMisurazione(descrizione="Tramezza sp 10 esistente camere",
                                    lung=5.0, larg=0.12, h_peso=3.3),
                    RigaMisurazione(descrizione="Tramezza sp 25 esistente bagno e antibagno",
                                    lung=2.35, larg=0.25, h_peso=3.3),
                    RigaMisurazione(descrizione="Tramezza sp 25 esistente bagno e antibagno",
                                    lung=1.88, larg=0.18, h_peso=3.3),
                    RigaMisurazione(descrizione="Nuova apertura tra soggiorno e notte",
                                    lung=1.0, larg=0.18, h_peso=2.2),
                    RigaMisurazione(descrizione="Pavimento bagno e antibagno esistente",
                                    lung=7.45, larg=0.25),
                ],
            ),
        ],
    )


def _carica_xlsx(preventivo: Preventivo) -> openpyxl.Workbook:
    """Genera l'xlsx e lo carica con openpyxl per ispezione."""
    return openpyxl.load_workbook(io.BytesIO(genera_xlsx(preventivo)))


# ---------------------------------------------------------------------------
# Struttura del foglio
# ---------------------------------------------------------------------------

class TestStrutturaFoglio:
    def setup_method(self):
        self.wb = _carica_xlsx(_preventivo_di_test())
        self.ws = self.wb["Computo metrico"]

    def test_fogli_presenti(self):
        assert "Computo metrico" in self.wb.sheetnames
        assert "Dati" in self.wb.sheetnames

    def test_intestazione_riga2(self):
        ws = self.ws
        assert ws.cell(2, 2).value == "Nr. Ord."
        assert ws.cell(2, 3).value == "TARIFFA"
        assert ws.cell(2, 4).value == "DESIGNAZIONE DEI LAVORI"
        assert ws.cell(2, 9).value == "Quantità"

    def test_intestazione_riga3(self):
        ws = self.ws
        assert ws.cell(3, 5).value == "Par.ug"
        assert ws.cell(3, 6).value == "Lung."
        assert ws.cell(3, 7).value == "Larg."
        assert ws.cell(3, 8).value == "H/peso"
        assert ws.cell(3, 10).value == "unitario"
        assert ws.cell(3, 11).value == "TOTALE"
        assert ws.cell(3, 12).value == "ClDes"
        assert ws.cell(3, 13).value == "ClQT"
        assert ws.cell(3, 14).value == "Linha"

    def test_colonne_nascoste(self):
        ws = self.ws
        assert ws.column_dimensions["L"].hidden is True
        assert ws.column_dimensions["M"].hidden is True
        assert ws.column_dimensions["N"].hidden is True


# ---------------------------------------------------------------------------
# Struttura voce 1 (1 misurazione)
# ---------------------------------------------------------------------------

class TestVoce1:
    def setup_method(self):
        self.ws = _carica_xlsx(_preventivo_di_test())["Computo metrico"]

    def test_intestazione(self):
        r = PRIMA_RIGA_DATI
        ws = self.ws
        assert ws.cell(r, 2).value == "1"
        assert ws.cell(r, 3).value == "001"
        assert "Approntamento" in ws.cell(r, 4).value
        # I e K nelle righe placeholder sono celle vuote (nessuna formula invalida)
        assert ws.cell(r, 9).value is None
        assert ws.cell(r, 11).value is None

    def test_label_misurazioni(self):
        assert self.ws.cell(PRIMA_RIGA_DATI + 1, 4).value == "M I S U R A Z I O N I:"

    def test_riga_misurazione(self):
        r = PRIMA_RIGA_DATI + 2
        assert self.ws.cell(r, 5).value == 1.0
        assert self.ws.cell(r, 9).value == pytest.approx(1.0)

    def test_marker_3prime(self):
        r = PRIMA_RIGA_DATI + 3
        assert self.ws.cell(r, 14).value == "3'"

    def test_sommano(self):
        r = PRIMA_RIGA_DATI + 4
        assert self.ws.cell(r, 4).value == "SOMMANO a corpo"
        assert self.ws.cell(r, 9).value == pytest.approx(1.0)   # quantita_x = par_ug = 1
        assert self.ws.cell(r, 11).value == pytest.approx(0.0)  # 1.0 * prezzo 0 = 0

    def test_separatore(self):
        r = PRIMA_RIGA_DATI + 5
        # Riga separatore: I e K sono celle vuote
        assert self.ws.cell(r, 9).value is None
        assert self.ws.cell(r, 11).value is None


# ---------------------------------------------------------------------------
# Struttura voce 3 (6 misurazioni)
# ---------------------------------------------------------------------------

class TestVoce3:
    RIGA_VOCE3 = 16  # voce 1 e 2 occupano 6 righe ciascuna → voce 3 inizia a riga 16

    def setup_method(self):
        self.ws = _carica_xlsx(_preventivo_di_test())["Computo metrico"]

    def test_intestazione(self):
        r = self.RIGA_VOCE3
        assert self.ws.cell(r, 2).value == "3"
        assert self.ws.cell(r, 3).value == "003"

    def test_sei_righe_misurazioni(self):
        """Tutte e 6 le misurazioni devono avere un valore numerico positivo in colonna I."""
        prima_riga_mis = self.RIGA_VOCE3 + 2
        righe_con_valore = sum(
            1 for offset in range(6)
            if isinstance(self.ws.cell(prima_riga_mis + offset, 9).value, (int, float))
            and self.ws.cell(prima_riga_mis + offset, 9).value > 0
        )
        assert righe_con_valore == 6

    def test_sommano_voce3_quantita(self):
        """Il SOMMANO deve riportare la somma corretta delle 6 misurazioni."""
        riga_marker = self.RIGA_VOCE3 + 2 + 6
        riga_sommano = riga_marker + 1
        valore = self.ws.cell(riga_sommano, 9).value
        assert isinstance(valore, (int, float))
        assert valore == pytest.approx(10.07, abs=0.01)


# ---------------------------------------------------------------------------
# Footer
# ---------------------------------------------------------------------------

class TestFooter:
    def setup_method(self):
        self.ws = _carica_xlsx(_preventivo_di_test())["Computo metrico"]

    def _trova_riga(self, testo_colonna_d: str) -> int:
        for riga in self.ws.iter_rows(min_row=PRIMA_RIGA_DATI):
            if riga[3].value == testo_colonna_d:
                return riga[0].row
        return None

    def test_totale_euro(self):
        riga = self._trova_riga("TOTALE euro")
        assert riga is not None
        valore_k = self.ws.cell(riga, 11).value
        assert isinstance(valore_k, (int, float))   # valore numerico, non formula
        assert self.ws.cell(riga, 12).value == "0"

    def test_aggiunge_nuova_voce(self):
        riga = self._trova_riga("AGGIUNGE NUOVA VOCE")
        assert riga is not None
        assert self.ws.cell(riga, 12).value == "0"

    def test_firma_primus(self):
        trovato = any(
            riga[1].value and "PriMus" in str(riga[1].value)
            for riga in self.ws.iter_rows(min_row=PRIMA_RIGA_DATI)
        )
        assert trovato


# ---------------------------------------------------------------------------
# Calcoli misurazioni
# ---------------------------------------------------------------------------

class TestCalcoliMisurazioni:
    def test_prodotto_tre_fattori(self):
        m = RigaMisurazione(lung=4.2, larg=0.2, h_peso=3.3)
        assert m.quantita == pytest.approx(2.77, abs=0.01)

    def test_un_solo_fattore(self):
        assert RigaMisurazione(par_ug=1.0).quantita == 1.0

    def test_due_fattori(self):
        assert RigaMisurazione(lung=5.0, h_peso=3.0).quantita == pytest.approx(15.0)

    def test_riga_vuota_restituisce_zero(self):
        assert RigaMisurazione().quantita == 0.0
        assert RigaMisurazione(descrizione="etichetta").quantita == 0.0

    def test_is_empty(self):
        assert RigaMisurazione().is_empty is True
        assert RigaMisurazione(descrizione="").is_empty is True
        assert RigaMisurazione(lung=4.2).is_empty is False
        assert RigaMisurazione(descrizione="nota", lung=4.2).is_empty is False

    def test_riga_vuota_non_altera_quantita_x(self):
        """Una riga vuota aggiunta dall'UI non deve influenzare la somma delle quantità."""
        voce = Voce(
            codice="001", descrizione="test", um="mc",
            misurazioni=[
                RigaMisurazione(lung=4.2, larg=0.2, h_peso=3.3),   # 2.77
                RigaMisurazione(lung=5.0, larg=0.12, h_peso=3.3),  # 1.98
                RigaMisurazione(),                                   # riga vuota UI
            ],
        )
        assert quantita_x(voce) == pytest.approx(4.75, abs=0.01)

    def test_riga_vuota_filtrata_nel_xlsx(self):
        """La riga vuota non deve influenzare la quantità nel SOMMANO."""
        prev = Preventivo(ricarica_default=0.0, voci=[
            Voce(
                codice="001", descrizione="Test", um="mc",
                misurazioni=[
                    RigaMisurazione(lung=4.2, larg=0.2, h_peso=3.3),   # 2.77
                    RigaMisurazione(lung=5.0, larg=0.12, h_peso=3.3),  # 1.98
                    RigaMisurazione(),  # riga vuota → ignorata
                ],
                costi=[RigaCosto(descrizione="op", um="ore", quantita=1, costo_unitario=1)],
            )
        ])
        ws = openpyxl.load_workbook(io.BytesIO(genera_xlsx(prev)))["Computo metrico"]
        # SOMMANO è a PRIMA_RIGA_DATI + 5 (header + label + 2 mis + marker + sommano)
        riga_sommano = PRIMA_RIGA_DATI + 5
        assert ws.cell(riga_sommano, 9).value == pytest.approx(4.75, abs=0.01)


# ---------------------------------------------------------------------------
# Quantità manuale
# ---------------------------------------------------------------------------

class TestQuantitaManuale:
    def test_export_scrive_par_ug(self):
        """Voce senza misurazioni: quantita_manuale deve finire in E (par_ug)."""
        prev = Preventivo(ricarica_default=0.0, voci=[
            Voce(
                codice="001", descrizione="A corpo", um="a corpo",
                quantita_manuale=5.0,
                misurazioni=[],
                costi=[RigaCosto(descrizione="op", um="ore", quantita=1, costo_unitario=100)],
            )
        ])
        ws = openpyxl.load_workbook(io.BytesIO(genera_xlsx(prev)))["Computo metrico"]
        riga_mis = PRIMA_RIGA_DATI + 2
        assert ws.cell(riga_mis, 5).value == 5.0

    def test_calcolo_con_quantita_manuale(self):
        """quantita_x usa quantita_manuale quando non ci sono misurazioni."""
        preventivo = Preventivo(ricarica_default=0.2, voci=[])
        voce = Voce(
            codice="001", descrizione="A corpo", um="a corpo",
            quantita_manuale=1.0,
            misurazioni=[],
            costi=[RigaCosto(descrizione="op", um="ore", quantita=2, costo_unitario=20)],
        )
        preventivo.voci.append(voce)
        assert quantita_x(voce) == 1.0
        assert prezzo_per_um(voce, preventivo) == pytest.approx(48.0)

    def test_calcolo_completo_voce(self):
        from cbd_preventivi.calcoli import totale_costi, costo_per_um

        preventivo = Preventivo(ricarica_default=0.20, voci=[])
        voce = Voce(
            codice="001", descrizione="Test", um="mc",
            misurazioni=[RigaMisurazione(lung=5.0, larg=2.0, h_peso=3.0)],  # 30 mc
            costi=[
                RigaCosto(descrizione="Operaio",   um="ore", quantita=10, costo_unitario=31.0),
                RigaCosto(descrizione="Autocarro", um="ore", quantita=5,  costo_unitario=65.0),
            ],
        )
        preventivo.voci.append(voce)
        assert quantita_x(voce) == pytest.approx(30.0)
        assert totale_costi(voce) == pytest.approx(635.0)
        assert costo_per_um(voce) == pytest.approx(635 / 30, rel=1e-4)
        assert prezzo_per_um(voce, preventivo) == pytest.approx((635 / 30) * 1.20, rel=1e-4)
